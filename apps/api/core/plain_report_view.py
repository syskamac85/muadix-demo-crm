"""Plain Django view for contact report - no DRF dependencies."""
import logging
from collections import defaultdict
from datetime import timedelta
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Count, Exists, OuterRef, Subquery
from django.views.decorators.http import require_http_methods
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.exceptions import AuthenticationFailed

from accounts.models import UserRole
from .models import Client, CallRecord
from .views import _build_pdf_response
from .services.contact_plan import PlanComputationContext, compute_due_date

logger = logging.getLogger(__name__)

User = get_user_model()


@require_http_methods(["GET"])
def plain_completed_export(request):
    """Export completed contacts for a specific date - plain Django view."""
    logger.info("=== plain_completed_export CALLED ===")
    
    # Manually authenticate using JWT
    auth = JWTAuthentication()
    try:
        user_auth = auth.authenticate(request)
        if user_auth is not None:
            request.user = user_auth[0]
        else:
            logger.warning("JWT authentication returned None")
            if not request.user.is_authenticated:
                return JsonResponse({'detail': 'Authentication required.'}, status=401)
    except AuthenticationFailed as e:
        logger.warning("JWT authentication failed: %s", e)
        return JsonResponse({'detail': str(e)}, status=401)
    except Exception as e:
        logger.error("JWT authentication error: %s", e)
        if not request.user.is_authenticated:
            return JsonResponse({'detail': 'Authentication required.'}, status=401)
    
    logger.info("User: %s", request.user)
    logger.info("GET params: %s", dict(request.GET))
    
    # Check authentication
    if not request.user.is_authenticated:
        return JsonResponse({'detail': 'Authentication required.'}, status=401)
    
    date_param = request.GET.get('date')
    export_format = request.GET.get('format', 'pdf').lower()
    
    target_date = parse_date(date_param) if date_param else timezone.localdate()
    
    # Get completed contacts for the date
    user = request.user
    queryset = CallRecord.objects.select_related('client', 'handler').filter(contact_date=target_date)
    
    # Apply role-based filtering
    if user.role == UserRole.ADMIN:
        pass  # Admin sees all
    elif user.role == UserRole.MANAGER:
        if user.tenant_id:
            queryset = queryset.filter(tenant=user.tenant)
    elif user.role == UserRole.REP:
        queryset = queryset.filter(handler=user)
    else:
        queryset = queryset.none()
    
    # Prepare data
    headers = ["Data", "Klient", "NIP", "Miasto", "Handlowiec", "Wynik", "Komentarz"]
    rows = []
    
    for record in queryset:
        handler = record.handler
        handler_name = f"{handler.first_name} {handler.last_name}".strip() if handler else ''
        if not handler_name and handler:
            handler_name = handler.username or ''
        rows.append([
            record.contact_date.isoformat() if record.contact_date else '',
            getattr(record.client, 'name', '') or '',
            getattr(record.client, 'nip', '') or '',
            getattr(record.client, 'city', '') or '',
            handler_name,
            record.outcome or '',
            record.current_comment or '',
        ])
    
    logger.info("plain_completed_export generating %s with %s rows", export_format.upper(), len(rows))
    
    if export_format == 'pdf':
        filename = f"wykonane_kontakty_{target_date.isoformat()}.pdf"
        title = f"Wykonane kontakty – {target_date.isoformat()}"
        return _build_pdf_response(
            filename,
            title,
            headers,
            rows,
            landscape_mode=True,
            column_widths=[1.0, 2.5, 1.2, 1.3, 1.5, 1.2, 3.0],
        )
    else:  # xlsx
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Wykonane kontakty"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"wykonane_kontakty_{target_date.isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response


@require_http_methods(["GET"])
def plain_contact_stats_history(request):
    """Generate historical day-by-day stats (planned vs completed)."""
    logger.info("=== plain_contact_stats_history CALLED ===")

    auth = JWTAuthentication()
    try:
        user_auth = auth.authenticate(request)
        if user_auth is not None:
            request.user = user_auth[0]
        elif not request.user.is_authenticated:
            return JsonResponse({'detail': 'Authentication required.'}, status=401)
    except AuthenticationFailed as exc:
        return JsonResponse({'detail': str(exc)}, status=401)
    except Exception as exc:  # pragma: no cover - defensive
        logger.error("JWT authentication error: %s", exc)
        if not request.user.is_authenticated:
            return JsonResponse({'detail': 'Authentication required.'}, status=401)

    date_from_param = request.GET.get('date_from')
    date_to_param = request.GET.get('date_to')
    export_format = request.GET.get('format', 'xlsx').lower()

    date_from = parse_date(date_from_param) if date_from_param else timezone.localdate()
    date_to = parse_date(date_to_param) if date_to_param else date_from

    if not date_from or not date_to:
        return JsonResponse({'detail': 'Podaj poprawne daty.'}, status=400)
    if date_from > date_to:
        return JsonResponse({'detail': 'Data od nie może być późniejsza niż data do.'}, status=400)

    user = request.user
    queryset = Client.objects.select_related('tenant', 'salesman').filter(status=Client.Status.ACTIVE)
    queryset = _apply_role_filter_to_clients(queryset, user)

    latest_records = CallRecord.objects.filter(client=OuterRef('pk')).order_by('-contact_date', '-id')
    base_queryset = queryset.annotate(
        last_contact_date=Subquery(latest_records.values('contact_date')[:1]),
        recorded_next_contact=Subquery(latest_records.values('next_contact_at')[:1]),
    )
    clients = list(base_queryset)

    day_stats: dict = {}
    current = date_from
    while current <= date_to:
        day_stats[current] = {
            'scheduled': defaultdict(int),
            'completed': defaultdict(int),
        }
        current += timedelta(days=1)

    current = date_from
    while current <= date_to:
        stats_for_day = day_stats[current]
        for client in clients:
            cycle_days = _resolve_cycle_days_for_client(client)
            if not cycle_days:
                continue
            client_tenant = getattr(client, 'tenant', None)
            salesman = getattr(client, 'salesman', None)
            salesman_cycle_start = getattr(salesman, 'contact_cycle_start_date', None) if salesman else None
            tenant_cycle_start = getattr(client_tenant, 'contact_cycle_start_date', None)
            created_date = client.created_at.date() if getattr(client, 'created_at', None) else current
            context = PlanComputationContext(
                created_date=created_date,
                cycle_days=cycle_days,
                last_contact_date=getattr(client, 'last_contact_date', None),
                recorded_next_date=getattr(client, 'recorded_next_contact', None),
                cycle_start_date=salesman_cycle_start or tenant_cycle_start,
            )
            result = compute_due_date(context, current)
            if result and result.due_date == current:
                salesman_name = _format_salesman_local(salesman) or 'Brak handlowca'
                stats_for_day['scheduled'][salesman_name] += 1
        current += timedelta(days=1)

    completed_qs = CallRecord.objects.filter(contact_date__gte=date_from, contact_date__lte=date_to)
    completed_qs = _apply_role_filter_to_call_records(completed_qs.select_related('handler'), user)
    for record in completed_qs:
        record_date = record.contact_date
        stats_for_day = day_stats.get(record_date)
        if not stats_for_day:
            continue
        salesman_name = _format_salesman_local(getattr(record, 'handler', None)) or 'Brak handlowca'
        stats_for_day['completed'][salesman_name] += 1

    headers = ["Data", "Handlowiec", "Zaplanowane", "Wykonane", "Skuteczność"]
    rows = []
    for date_key in sorted(day_stats.keys()):
        stats_for_day = day_stats[date_key]
        salesman_names = sorted(set(stats_for_day['scheduled'].keys()) | set(stats_for_day['completed'].keys()))
        # Skip days with no activity
        if not salesman_names:
            continue
        for salesman_name in salesman_names:
            scheduled = stats_for_day['scheduled'].get(salesman_name, 0)
            completed = stats_for_day['completed'].get(salesman_name, 0)
            completion_rate = 0.0 if scheduled == 0 else (completed / scheduled) * 100
            rows.append([
                date_key.isoformat(),
                salesman_name,
                scheduled,
                completed,
                f"{completion_rate:.0f}%",
            ])

    if export_format == 'pdf':
        filename = f"historia_statystyk_{date_from.isoformat()}_{date_to.isoformat()}.pdf"
        title = f"Historia statystyk – {date_from.isoformat()} do {date_to.isoformat()}"
        return _build_pdf_response(
            filename,
            title,
            headers,
            rows,
            landscape_mode=False,
            column_widths=[1.2, 2.6, 1.2, 1.2, 1.2],
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statystyki"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historia_statystyk_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


# Shared helpers ------------------------------------------------------------


def _resolve_cycle_days_for_client(client):
    label = getattr(client, 'contact_days_label', '') or ''
    stripped = label.strip()
    if stripped:
        normalized = stripped.replace(',', '.').split()[0]
        try:
            value = float(normalized)
            if value > 0:
                return int(round(value))
        except ValueError:
            digits = ''.join(ch for ch in stripped if ch.isdigit())
            if digits:
                return int(digits)
    if getattr(client, 'contact_reminder_days', None) and client.contact_reminder_days > 0:
        return int(client.contact_reminder_days)
    return None


def _apply_role_filter_to_clients(queryset, user):
    if user.role == UserRole.ADMIN:
        return queryset
    if user.role == UserRole.MANAGER:
        if user.tenant_id:
            return queryset.filter(tenant=user.tenant)
        return queryset.none()
    if user.role == UserRole.REP:
        if user.tenant_id:
            return queryset.filter(tenant=user.tenant, salesman=user)
        return queryset.none()
    return queryset.none()


def _apply_role_filter_to_call_records(queryset, user):
    if user.role == UserRole.ADMIN:
        return queryset
    if user.role == UserRole.MANAGER:
        if user.tenant_id:
            return queryset.filter(tenant=user.tenant)
        return queryset.none()
    if user.role == UserRole.REP:
        if user.tenant_id:
            return queryset.filter(tenant=user.tenant, handler=user)
        return queryset.none()
    return queryset.none()


def _format_salesman_local(user):
    if not user:
        return None
    full_name = f"{getattr(user, 'first_name', '') or ''} {getattr(user, 'last_name', '') or ''}".strip()
    username = getattr(user, 'username', None)
    return full_name or username


@require_http_methods(["GET"])
def plain_contact_report(request):
    """Generate contact report in PDF or XLSX format - plain Django view."""
    logger.info("=== plain_contact_report CALLED ===")
    
    # Manually authenticate using JWT
    auth = JWTAuthentication()
    try:
        user_auth = auth.authenticate(request)
        if user_auth is not None:
            request.user = user_auth[0]
        else:
            logger.warning("JWT authentication returned None")
            if not request.user.is_authenticated:
                return JsonResponse({'detail': 'Authentication required.'}, status=401)
    except AuthenticationFailed as e:
        logger.warning("JWT authentication failed: %s", e)
        return JsonResponse({'detail': str(e)}, status=401)
    except Exception as e:
        logger.error("JWT authentication error: %s", e)
        if not request.user.is_authenticated:
            return JsonResponse({'detail': 'Authentication required.'}, status=401)
    
    logger.info("User: %s", request.user)
    logger.info("GET params: %s", dict(request.GET))
    
    # Check authentication
    if not request.user.is_authenticated:
        return JsonResponse({'detail': 'Authentication required.'}, status=401)
    
    date_from_param = request.GET.get('date_from')
    date_to_param = request.GET.get('date_to')
    export_format = request.GET.get('format', 'pdf').lower()

    date_from = parse_date(date_from_param) if date_from_param else timezone.localdate()
    date_to = parse_date(date_to_param) if date_to_param else timezone.localdate()

    logger.info(
        "plain_contact_report params: user=%s from=%s to=%s format=%s",
        request.user.id,
        date_from,
        date_to,
        export_format,
    )

    if date_from > date_to:
        return JsonResponse({'detail': 'Data od nie może być późniejsza niż data do.'}, status=400)

    # Get base queryset with role-based filtering
    user = request.user
    queryset = Client.objects.select_related('tenant', 'salesman').all()
    queryset = _apply_role_filter_to_clients(queryset, user)

    scheduled = []
    completed = []

    # Annotate with latest contacts
    latest_records = CallRecord.objects.filter(client=OuterRef('pk')).order_by('-contact_date', '-id')
    base_queryset = queryset.annotate(
        last_contact_date=Subquery(latest_records.values('contact_date')[:1]),
        recorded_next_contact=Subquery(latest_records.values('next_contact_at')[:1]),
    )

    # Get scheduled contacts for date range
    current = date_from
    
    while current <= date_to:
        completion_subquery = CallRecord.objects.filter(client=OuterRef('pk'), contact_date=current)
        qs = base_queryset.annotate(completed_on_selected=Exists(completion_subquery))
        
        for client in qs:
            cycle_days = _resolve_cycle_days_for_client(client)
            if not cycle_days:
                continue
            
            created_date = (client.created_at.date() if client.created_at else current)
            client_tenant = getattr(client, 'tenant', None)
            salesman = getattr(client, 'salesman', None)
            salesman_cycle_start = getattr(salesman, 'contact_cycle_start_date', None) if salesman else None
            tenant_cycle_start = getattr(client_tenant, 'contact_cycle_start_date', None)
            client_cycle_start = salesman_cycle_start or tenant_cycle_start
            
            context = PlanComputationContext(
                created_date=created_date,
                cycle_days=cycle_days,
                last_contact_date=getattr(client, 'last_contact_date', None),
                recorded_next_date=getattr(client, 'recorded_next_contact', None),
                cycle_start_date=client_cycle_start,
            )
            result = compute_due_date(context, current)
            if not result:
                continue
            
            # Format salesman name
            salesman_name = None
            if salesman:
                full_name = f"{salesman.first_name or ''} {salesman.last_name or ''}".strip()
                salesman_name = full_name or salesman.username
            
            entry = {
                'date': current.isoformat(),
                'name': client.name,
                'nip': client.nip,
                'city': client.city,
                'salesman_name': salesman_name,
                'cycle_days': cycle_days,
                'last_contact_date': context.last_contact_date.isoformat() if context.last_contact_date else None,
                'due_date': result.due_date.isoformat(),
                'completed_on_selected': bool(getattr(client, 'completed_on_selected', False)),
            }
            if entry['due_date'] == current.isoformat():
                scheduled.append(entry)
        
        current += timedelta(days=1)

    # Get completed contacts for date range
    completed_qs = CallRecord.objects.filter(contact_date__gte=date_from, contact_date__lte=date_to)
    completed_qs = _apply_role_filter_to_call_records(completed_qs, user)
    
    completed_qs = completed_qs.select_related('client', 'handler')
    
    for record in completed_qs:
        handler = record.handler
        handler_name = f"{handler.first_name} {handler.last_name}".strip() if handler else ''
        if not handler_name and handler:
            handler_name = handler.username or ''
        completed.append({
            'date': record.contact_date.isoformat(),
            'name': getattr(record.client, 'name', '') or '',
            'nip': getattr(record.client, 'nip', '') or '',
            'city': getattr(record.client, 'city', '') or '',
            'salesman_name': handler_name,
            'outcome': record.outcome or '',
            'current_comment': record.current_comment or '',
        })

    # Combine scheduled and completed into one row per client
    # Group by (date, client_name, nip)
    client_map = {}
    
    for entry in scheduled:
        key = (entry['date'], entry['name'], entry['nip'])
        if key not in client_map:
            client_map[key] = {
                'date': entry['date'],
                'name': entry['name'],
                'nip': entry['nip'],
                'city': entry['city'],
                'salesman_name': entry['salesman_name'] or '-',
                'planned': True,
                'cycle_days': entry['cycle_days'],
                'completed': False,
                'outcome': '',
                'comment': '',
            }
    
    for entry in completed:
        key = (entry['date'], entry['name'], entry['nip'])
        if key in client_map:
            # Client was both planned and completed
            client_map[key]['completed'] = True
            client_map[key]['outcome'] = entry['outcome']
            client_map[key]['comment'] = entry['current_comment']
        else:
            # Client was completed but not planned
            client_map[key] = {
                'date': entry['date'],
                'name': entry['name'],
                'nip': entry['nip'],
                'city': entry['city'],
                'salesman_name': entry['salesman_name'],
                'planned': False,
                'cycle_days': None,
                'completed': True,
                'outcome': entry['outcome'],
                'comment': entry['current_comment'],
            }
    
    # Generate file
    headers = [
        "Data",
        "Klient",
        "NIP",
        "Miasto",
        "Handlowiec",
        "Planowany",
        "Wykonany",
        "Szczegóły",
    ]
    rows = []
    for key in sorted(client_map.keys()):
        entry = client_map[key]
        
        # Status columns
        planned_status = "TAK" if entry['planned'] else "NIE"
        completed_status = "TAK" if entry['completed'] else "NIE"
        
        # Details column - only show outcome and comment
        if entry['completed']:
            details = f"{entry['outcome']} - {entry['comment']}" if entry['outcome'] or entry['comment'] else "-"
        else:
            details = "-"
        
        rows.append([
            entry['date'],
            entry['name'],
            entry['nip'],
            entry['city'],
            entry['salesman_name'],
            planned_status,
            completed_status,
            details,
        ])

    logger.info("plain_contact_report generating %s with %s rows", export_format.upper(), len(rows))

    if export_format == 'pdf':
        filename = f"raport_kontaktow_{date_from.isoformat()}_{date_to.isoformat()}.pdf"
        title = f"Raport kontaktów – {date_from.isoformat()} do {date_to.isoformat()}"
        return _build_pdf_response(
            filename,
            title,
            headers,
            rows,
            landscape_mode=True,
            column_widths=[1.0, 2.2, 1.0, 1.2, 1.4, 0.9, 0.9, 2.4],
        )
    else:  # xlsx
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Raport kontaktów"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"raport_kontaktow_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
