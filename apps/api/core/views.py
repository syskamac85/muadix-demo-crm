import logging
from datetime import datetime, date as date_cls, timedelta
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any, Optional, Sequence

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db.models import Exists, OuterRef, Subquery, Count, Q
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from rest_framework import mixins, permissions, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import UserRole

from .models import (
    AuditLog,
    BackupJob,
    CallRecord,
    Client,
    ClientDeletionRequest,
    Comment,
    ContactNextDateRequest,
    ImportJob,
    ImportRecord,
    RoutePlan,
    RouteStop,
    Task,
    TaskMessage,
    Tenant,
    Visit,
)
from .permissions import IsManagerOrAdmin
from .serializers import (
    AuditLogSerializer,
    BackupJobSerializer,
    CallRecordSerializer,
    ClientDeletionRequestSerializer,
    ClientSerializer,
    CommentSerializer,
    ContactNextDateRequestSerializer,
    ImportJobSerializer,
    ImportRecordSerializer,
    RoutePlanSerializer,
    RouteStopSerializer,
    TaskMessageSerializer,
    TaskSerializer,
    TenantSerializer,
    VisitSerializer,
)
from .services.contact_plan import PlanComputationContext, compute_due_date
from .services.working_days import shift_to_business_day
from .services.regon import RegonAPIError, lookup_client_by_nip
from .services.regon_scraper import lookup_client_by_nip_scraper
from .tasks import generate_pg_dump_bytes, process_import_job, run_backup_job
from .broadcast import broadcast_task_update, broadcast_task_message

User = get_user_model()
logger = logging.getLogger(__name__)

_PDF_FONT_NAME: Optional[str] = None


def _find_pdf_font_candidate() -> Optional[Path]:
    base_dir = getattr(settings, 'BASE_DIR', None)
    candidates: list[Path] = []
    if base_dir:
        candidates.extend([
            Path(base_dir) / 'fonts' / 'DejaVuSans.ttf',
            Path(base_dir) / 'assets' / 'fonts' / 'DejaVuSans.ttf',
        ])
    candidates.extend([
        Path('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
        Path('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'),
        Path('/usr/share/fonts/truetype/freefont/FreeSans.ttf'),
    ])
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def _ensure_pdf_font() -> str:
    global _PDF_FONT_NAME
    if _PDF_FONT_NAME:
        return _PDF_FONT_NAME
    candidate = _find_pdf_font_candidate()
    if candidate:
        font_name = 'DejaVuSans'
        pdfmetrics.registerFont(TTFont(font_name, str(candidate)))
        _PDF_FONT_NAME = font_name
    else:
        _PDF_FONT_NAME = 'Helvetica'
    return _PDF_FONT_NAME


def _resolve_column_widths(total_width: float, ratios: Optional[Sequence[float]]) -> Optional[list[float]]:
    if not ratios:
        return None
    total = sum(ratios)
    if not total:
        return None
    scale = total_width / total
    return [value * scale for value in ratios]


def _paragraphize(value: object, style) -> Paragraph:
    text = '' if value in (None, '') else str(value)
    safe_text = escape(text).replace('\n', '<br />')
    return Paragraph(safe_text, style)


def _build_pdf_response(
    filename: str,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    *,
    landscape_mode: bool = False,
    column_widths: Optional[Sequence[float]] = None,
) -> HttpResponse:
    buffer = BytesIO()
    font_name = _ensure_pdf_font()
    page_size = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        title=title,
        leftMargin=24,
        rightMargin=24,
        topMargin=32,
        bottomMargin=32,
    )
    styles = getSampleStyleSheet()
    title_style = styles['Heading2'].clone('PDFTitle')
    title_style.fontName = font_name
    title_style.spaceAfter = 12
    story = [Paragraph(title, title_style), Spacer(1, 12)]

    header_style = styles['BodyText'].clone('PDFHeader')
    header_style.fontName = font_name
    header_style.fontSize = 9
    header_style.leading = 12
    header_style.textColor = colors.whitesmoke

    body_style = styles['BodyText'].clone('PDFBody')
    body_style.fontName = font_name
    body_style.fontSize = 9
    body_style.leading = 12

    data: list[list[Paragraph]] = [
        [_paragraphize(header, header_style) for header in headers]
    ]
    for row in rows:
        data.append([_paragraphize(value, body_style) for value in row])

    table = Table(data, repeatRows=1, colWidths=_resolve_column_widths(doc.width, column_widths))
    table.hAlign = 'LEFT'
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _diff_dict(before: dict[str, Any], after: dict[str, Any]) -> dict[str, dict[str, Any]]:
    changes: dict[str, dict[str, Any]] = {}
    keys = set(before.keys()) | set(after.keys())
    for key in keys:
        if before.get(key) != after.get(key):
            changes[key] = {'from': before.get(key), 'to': after.get(key)}
    return changes


def _log_audit_event(*, tenant, actor, event_type: str, entity_type: str, entity_id: int, changes: Optional[dict[str, dict[str, Any]]] = None):
    AuditLog.objects.create(
        tenant=tenant,
        actor=actor,
        event_type=event_type,
        entity_type=entity_type,
        entity_id=entity_id,
        changes=changes or {},
    )


def _serialize_client(client: Client) -> dict[str, Any]:
    return {
        'name': client.name,
        'nip': client.nip,
        'city': client.city,
        'postal_code': client.postal_code,
        'street': client.street,
        'classification': client.classification,
        'salesman': client.salesman_id,
        'phone': client.phone,
        'email': client.email,
        'location_name': client.location_name,
        'contact_reminder_days': client.contact_reminder_days,
        'contact_days_label': client.contact_days_label,
        'status': client.status,
        'deleted_at': client.deleted_at.isoformat() if client.deleted_at else None,
    }


def _serialize_task(task: Task) -> dict[str, Any]:
    return {
        'title': task.title,
        'description': task.description,
        'status': task.status,
        'due_date': task.due_date.isoformat() if task.due_date else None,
        'assigned_to': task.assigned_to_id,
        'client': task.client_id,
        'completed_at': task.completed_at.isoformat() if task.completed_at else None,
        'completed_by': task.completed_by_id,
        'created_by': task.created_by_id,
    }


def _serialize_route(route: RoutePlan) -> dict[str, Any]:
    stops = [
        {
            'client': stop.client_id,
            'order': stop.order,
            'drive_minutes': stop.drive_minutes,
            'visit_minutes': stop.visit_minutes,
        }
        for stop in route.stops.all().order_by('order')
    ]
    return {
        'date': route.date.isoformat() if route.date else None,
        'owner': route.owner_id,
        'total_drive_minutes': route.total_drive_minutes,
        'total_visit_minutes': route.total_visit_minutes,
        'shared_with_manager': route.shared_with_manager,
        'approval_status': route.approval_status,
        'approved_by': route.approved_by_id,
        'approved_at': route.approved_at.isoformat() if route.approved_at else None,
        'stops': stops,
    }


def _serialize_visit(visit: Visit) -> dict[str, Any]:
    return {
        'client': visit.client_id,
        'salesman': visit.salesman_id,
        'planned_at': visit.planned_at.isoformat() if visit.planned_at else None,
        'duration_minutes': visit.duration_minutes,
        'status': visit.status,
        'comment': visit.comment,
        'location_name': visit.location_name,
    }


def _serialize_call_record(record: CallRecord) -> dict[str, Any]:
    client = getattr(record, 'client', None)
    cycle_days = _resolve_client_cycle_days(client) if client else None
    return {
        'client': record.client_id,
        'handler': record.handler_id,
        'contact_date': record.contact_date.isoformat() if record.contact_date else None,
        'next_contact_at': record.next_contact_at.isoformat() if record.next_contact_at else None,
        'outcome': record.outcome,
        'current_comment': record.current_comment,
        'previous_comment': record.previous_comment,
        'cycle_days': cycle_days,
    }


def _format_created_changes(snapshot: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {key: {'from': None, 'to': value} for key, value in snapshot.items()}


def _format_deleted_changes(snapshot: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {key: {'from': value, 'to': None} for key, value in snapshot.items()}


def _resolve_client_cycle_days(client) -> int | None:
    """Zwraca liczbę dni cyklu klienta. Źródłem prawdy jest contact_days_label;
    contact_reminder_days używane tylko gdy label jest pusty."""
    label = getattr(client, 'contact_days_label', '') or ''
    stripped = label.strip()
    if stripped:
        normalized = stripped.replace(',', '.').split()[0]
        try:
            value = float(normalized)
            if value > 0:
                return int(round(value))
        except ValueError:
            digits = ''.join(ch for ch in stripped if ch.isdigit())
            if digits:
                return int(digits)
    if client.contact_reminder_days and client.contact_reminder_days > 0:
        return int(client.contact_reminder_days)
    return None


def _create_contact_next_date_request_if_needed(record: CallRecord, cycle_days: int | None, reason: str = '') -> bool:
    """Tworzy ContactNextDateRequest gdy next_contact_at > 2 × cycle_days.
    Zwraca True jeśli wniosek został utworzony."""
    if not cycle_days or not record.next_contact_at or not record.contact_date:
        return False
    proposed_days = (record.next_contact_at - record.contact_date).days
    if proposed_days <= 2 * cycle_days:
        return False
    ContactNextDateRequest.objects.get_or_create(
        call_record=record,
        defaults=dict(
            tenant=record.tenant,
            client=record.client,
            requested_by=record.handler,
            cycle_days=cycle_days,
            proposed_days=proposed_days,
            status=ContactNextDateRequest.Status.PENDING,
            reason=reason,
        ),
    )
    return True


class TenantScopedViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def filter_queryset_by_role(self, queryset):
        user = self.request.user
        if not user or not user.is_authenticated:
            return queryset.none()

        if user.role == UserRole.ADMIN:
            return queryset

        tenant = getattr(user, 'tenant', None)
        if tenant:
            return queryset.filter(tenant=tenant)
        return queryset.none()

    @action(detail=False, methods=['get'], url_path='contact-plan')
    def contact_plan(self, request):
        selected_param = request.query_params.get('date')
        selected_date = parse_date(selected_param) if selected_param else None
        if not selected_date:
            selected_date = timezone.localdate()

        export_flag = request.query_params.get('export')

        queryset = self._annotate_with_latest_contacts(self.filter_queryset(self.get_queryset()))
        completion_subquery = CallRecord.objects.filter(client=OuterRef('pk'), contact_date=selected_date)
        queryset = queryset.annotate(completed_on_selected=Exists(completion_subquery))
        entries: list[dict] = []
        tenant = getattr(request.user, 'tenant', None)
        if not tenant:
            tenant = Tenant.objects.first()
        tenant_cycle_start = getattr(tenant, 'contact_cycle_start_date', None)
        if tenant and tenant_cycle_start is None:
            tenant_cycle_start = timezone.localdate()
            tenant.contact_cycle_start_date = tenant_cycle_start
            tenant.save(update_fields=['contact_cycle_start_date'])

        for client in queryset:
            cycle_days = self._resolve_cycle_days(client)
            if not cycle_days:
                continue

            created_date = (client.created_at.date() if client.created_at else selected_date)
            client_tenant = getattr(client, 'tenant', None)
            salesman = getattr(client, 'salesman', None)
            salesman_cycle_start = getattr(salesman, 'contact_cycle_start_date', None) if salesman else None
            tenant_cycle_value = getattr(client_tenant, 'contact_cycle_start_date', None) or tenant_cycle_start
            client_cycle_start = salesman_cycle_start or tenant_cycle_value
            context = PlanComputationContext(
                created_date=created_date,
                cycle_days=cycle_days,
                last_contact_date=getattr(client, 'last_contact_date', None),
                recorded_next_date=getattr(client, 'recorded_next_contact', None),
                cycle_start_date=client_cycle_start,
            )
            result = compute_due_date(context, selected_date)
            if not result:
                continue

            entry = {
                'client_id': client.id,
                'name': client.name,
                'nip': client.nip,
                'phone': client.phone,
                'email': client.email,
                'city': client.city,
                'salesman_name': self._format_salesman(salesman),
                'salesman_id': getattr(salesman, 'id', None),
                'cycle_days': cycle_days,
                'contact_days_label': client.contact_days_label,
                'contact_cycle_start_date': client_cycle_start.isoformat() if client_cycle_start else None,
                'last_contact_date': context.last_contact_date.isoformat() if context.last_contact_date else None,
                'recorded_next_contact': context.recorded_next_date.isoformat() if context.recorded_next_date else None,
                'due_date': result.due_date.isoformat(),
                'raw_due_date': result.raw_due_date.isoformat(),
                'previous_due_date': result.previous_due_date.isoformat() if result.previous_due_date else None,
                'is_due_on_selected': result.due_date == selected_date,
                'completed_on_selected': bool(getattr(client, 'completed_on_selected', False)),
            }
            entries.append(entry)

        entries.sort(key=lambda item: (item['due_date'], item['name']))
        due_on_selected = [entry for entry in entries if entry['is_due_on_selected']]
        next_available_date = entries[0]['due_date'] if entries else None

        if export_flag in {'xlsx', 'pdf'}:
            rows = [
                [
                    entry['due_date'],
                    entry['name'],
                    entry['nip'],
                    entry['city'],
                    entry['salesman_name'] or '-',
                    entry['cycle_days'],
                    entry['last_contact_date'] or '',
                    entry['previous_due_date'] or '',
                ]
                for entry in due_on_selected
            ]
            if export_flag == 'pdf':
                filename = f"plan_kontaktow_{selected_date.isoformat()}.pdf"
                headers = [
                    "Data",
                    "Klient",
                    "NIP",
                    "Miasto",
                    "Handlowiec",
                    "Cykl (dni)",
                    "Ostatni kontakt",
                    "Poprzedni termin",
                ]
                return _build_pdf_response(
                    filename,
                    f"Plan kontaktów – {selected_date.isoformat()}",
                    headers,
                    rows,
                    landscape_mode=True,
                    column_widths=[1.2, 2.6, 1.2, 1.4, 1.6, 1.1, 1.4, 1.4],
                )
            return self._export_plan_excel(selected_date, due_on_selected)

        return Response(
            {
                'selected_date': selected_date.isoformat(),
                'due_on_selected': due_on_selected,
                'entries': entries,
                'counts': {
                    'due_on_selected': len(due_on_selected),
                    'total_schedulable': len(entries),
                },
                'global_cycle_start_date': tenant_cycle_start.isoformat() if tenant_cycle_start else None,
                'next_available_date': next_available_date,
            }
        )

    @action(detail=False, methods=['get'], url_path='contact-stats')
    def contact_stats(self, request):
        date_param = request.query_params.get('date')
        target_date = parse_date(date_param) if date_param else timezone.localdate()
        user = request.user

        queryset = self._annotate_with_latest_contacts(self.filter_queryset(self.get_queryset()))
        completion_subquery = CallRecord.objects.filter(client=OuterRef('pk'), contact_date=target_date)
        queryset = queryset.annotate(completed_on_selected=Exists(completion_subquery))

        scheduled_by_salesman = {}
        for client in queryset:
            cycle_days = self._resolve_cycle_days(client)
            if not cycle_days:
                continue
            created_date = (client.created_at.date() if client.created_at else target_date)
            client_tenant = getattr(client, 'tenant', None)
            salesman = getattr(client, 'salesman', None)
            salesman_cycle_start = getattr(salesman, 'contact_cycle_start_date', None) if salesman else None
            tenant_cycle_start = getattr(client_tenant, 'contact_cycle_start_date', None)
            client_cycle_start = salesman_cycle_start or tenant_cycle_start
            context = PlanComputationContext(
                created_date=created_date,
                cycle_days=cycle_days,
                last_contact_date=getattr(client, 'last_contact_date', None),
                recorded_next_date=getattr(client, 'recorded_next_contact', None),
                cycle_start_date=client_cycle_start,
            )
            result = compute_due_date(context, target_date)
            if result and result.due_date == target_date:
                salesman_name = self._format_salesman(salesman) or 'Brak handlowca'
                if user.role == UserRole.REP and salesman_name != self._format_salesman(user):
                    continue
                scheduled_by_salesman[salesman_name] = scheduled_by_salesman.get(salesman_name, 0) + 1

        completed_qs = self.filter_queryset(CallRecord.objects.filter(contact_date=target_date))
        completed_qs = completed_qs.select_related('handler')
        completed_by_salesman = {}
        for record in completed_qs:
            handler = getattr(record, 'handler', None)
            salesman_name = self._format_salesman(handler) or 'Brak handlowca'
            if user.role == UserRole.REP:
                if handler is None or handler.id != user.id:
                    continue
            completed_by_salesman[salesman_name] = completed_by_salesman.get(salesman_name, 0) + 1

        all_salesmen = set(scheduled_by_salesman.keys()) | set(completed_by_salesman.keys())
        stats_by_salesman = []
        for salesman_name in sorted(all_salesmen):
            stats_by_salesman.append({
                'salesman': salesman_name,
                'scheduled': scheduled_by_salesman.get(salesman_name, 0),
                'completed': completed_by_salesman.get(salesman_name, 0),
            })

        return Response({
            'by_salesman': stats_by_salesman,
            'total_scheduled': sum(scheduled_by_salesman.values()),
            'total_completed': sum(completed_by_salesman.values()),
        })

        if user.role == UserRole.MANAGER:
            if user.tenant_id:
                return queryset.filter(tenant=user.tenant)
            return queryset

        if user.role == UserRole.REP:
            if user.tenant_id:
                return queryset.filter(tenant=user.tenant)
            return queryset.none()

        return queryset.none()

    def get_queryset(self):
        qs = super().get_queryset()
        return self.filter_queryset_by_role(qs)


class TenantViewSet(TenantScopedViewSet):
    queryset = Tenant.objects.all()
    serializer_class = TenantSerializer


class ClientViewSet(TenantScopedViewSet):
    queryset = Client.objects.select_related('tenant', 'salesman').all()
    serializer_class = ClientSerializer
    lookup_value_regex = r'[0-9]+'
    permission_classes = [permissions.IsAuthenticated]

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role == UserRole.REP:
            return queryset.filter(salesman=user)
        return queryset

    def get_queryset(self):
        queryset = super().get_queryset()
        include_deleted = self.request.query_params.get('include_deleted')
        status_param = self.request.query_params.get('status')
        action = getattr(self, 'action', None)

        if action == 'restore':
            queryset = queryset.filter(status=Client.Status.DELETED)
        elif status_param:
            queryset = queryset.filter(status=status_param)
        elif include_deleted not in {'1', 'true', 'True'}:
            queryset = queryset.filter(status=Client.Status.ACTIVE)

        salesman_id = self.request.query_params.get('salesman')
        if salesman_id:
            try:
                queryset = queryset.filter(salesman_id=int(salesman_id))
            except (TypeError, ValueError):
                queryset = queryset.none()
        return queryset

    def perform_create(self, serializer):
        tenant = getattr(self.request.user, 'tenant', None)
        tenant_id = self.request.data.get('tenant') or self.request.data.get('tenant_id')
        if not tenant and tenant_id:
            try:
                tenant = Tenant.objects.filter(pk=int(tenant_id)).first()
            except (TypeError, ValueError):
                tenant = None
        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant': 'Brak przypisanego tenant.'})
        client = serializer.save(tenant=tenant)
        _log_audit_event(
            tenant=client.tenant,
            actor=self.request.user,
            event_type='client.created',
            entity_type='client',
            entity_id=client.id,
            changes=_format_created_changes(_serialize_client(client)),
        )

    def perform_update(self, serializer):
        instance = serializer.instance
        before = _serialize_client(instance)
        client = serializer.save()
        after = _serialize_client(client)
        changes = _diff_dict(before, after)
        if changes:
            _log_audit_event(
                tenant=client.tenant,
                actor=self.request.user,
                event_type='client.updated',
                entity_type='client',
                entity_id=client.id,
                changes=changes,
            )

    def perform_destroy(self, instance):
        snapshot = _serialize_client(instance)
        tenant = instance.tenant
        entity_id = instance.id
        response = super().perform_destroy(instance)
        _log_audit_event(
            tenant=tenant,
            actor=self.request.user,
            event_type='client.deleted',
            entity_type='client',
            entity_id=entity_id,
            changes=_format_deleted_changes(snapshot),
        )
        return response

    @action(detail=True, methods=['post'], url_path='request-deletion')
    def request_deletion(self, request, pk=None):
        client = self.get_object()
        if client.status == Client.Status.DELETED:
            return Response({'detail': 'Klient jest już oznaczony jako usunięty.'}, status=status.HTTP_400_BAD_REQUEST)

        pending_exists = ClientDeletionRequest.objects.filter(
            client=client,
            status=ClientDeletionRequest.Status.PENDING,
        ).exists()
        if pending_exists:
            return Response({'detail': 'Wniosek o usunięcie dla tego klienta jest już w toku.'}, status=status.HTTP_400_BAD_REQUEST)

        reason = request.data.get('reason', '')
        deletion_request = ClientDeletionRequest.objects.create(
            tenant=client.tenant,
            client=client,
            requested_by=request.user,
            reason=reason,
        )

        _log_audit_event(
            tenant=client.tenant,
            actor=request.user,
            event_type='client.deletion_requested',
            entity_type='client',
            entity_id=client.id,
            changes={
                'deletion_request': {
                    'from': None,
                    'to': str(deletion_request.id),
                }
            },
        )

        serializer = ClientDeletionRequestSerializer(deletion_request, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def restore(self, request, pk=None):
        client = self.get_object()
        if client.status != Client.Status.DELETED:
            return Response({'detail': 'Klient jest już aktywny.'}, status=status.HTTP_400_BAD_REQUEST)

        client.status = Client.Status.ACTIVE
        client.deleted_at = None
        client.updated_at = timezone.now()
        client.save(update_fields=['status', 'deleted_at', 'updated_at'])

        _log_audit_event(
            tenant=client.tenant,
            actor=request.user,
            event_type='client.restored',
            entity_type='client',
            entity_id=client.id,
            changes={'status': {'from': Client.Status.DELETED, 'to': Client.Status.ACTIVE}},
        )

        serializer = ClientSerializer(client, context={'request': request})
        return Response(serializer.data)

    @staticmethod
    def _parse_cycle_from_label(label: str | None) -> int | None:
        if not label:
            return None
        stripped = label.strip()
        if not stripped:
            return None
        normalized = stripped.replace(',', '.').split()[0]
        try:
            value = float(normalized)
            if value > 0:
                return int(round(value))
        except ValueError:
            digits = ''.join(ch for ch in stripped if ch.isdigit())
            if digits:
                return int(digits)
        return None

    @staticmethod
    def _format_salesman(user):
        if not user:
            return None
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        return full_name or user.username

    @action(detail=False, methods=['get'], url_path='export-excel', permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def export_excel(self, request):
        from openpyxl import Workbook
        from django.http import HttpResponse

        queryset = self.filter_queryset_by_role(
            Client.objects.select_related('tenant', 'salesman')
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Klienci"

        headers = [
            "Nazwa kontrahenta",
            "NIP",
            "Miasto",
            "Kod",
            "Ulica, nr lokalu",
            "Nazwa klasyfikacji",
            "Handlowiec",
        ]
        # Allow extending the sheet with extra columns while keeping the template-compatible prefix.
        headers.extend([
            "Typ klienta",
            "Dni_kontakt",
            "Telefon",
            "E-mail",
            "Imię handlowca",
            "Nazwisko handlowca",
            "Długość geograficzna",
            "Szerokość geograficzna",
            "Status",
        ])
        sheet.append(headers)

        def resolve_contact_cycle_value(client_obj):
            if client_obj.contact_days_label:
                return client_obj.contact_days_label
            if client_obj.contact_reminder_days:
                return str(client_obj.contact_reminder_days)
            return ""

        for client in queryset:
            salesman_name = ""
            if client.salesman:
                salesman_name = client.salesman.username

            status_label = "aktywny" if client.status == Client.Status.ACTIVE else "usunięty"
            sheet.append([
                client.name,
                client.nip or "",
                client.city or "",
                client.postal_code or "",
                client.street or "",
                client.classification or "",
                salesman_name,
                client.type or "",
                resolve_contact_cycle_value(client),
                client.phone or "",
                client.email or "",
                (client.salesman.first_name if client.salesman else "") or "",
                (client.salesman.last_name if client.salesman else "") or "",
                client.longitude or "",
                client.latitude or "",
                status_label,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"baza_klientow_{timezone.localdate().isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['delete'], url_path='purge', permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def purge_clients(self, request):
        tenant_param = request.query_params.get('tenant')
        queryset = self.get_queryset()
        if tenant_param:
            try:
                tenant_id = int(tenant_param)
            except (TypeError, ValueError):
                return Response({'detail': 'Nieprawidłowy identyfikator tenanta.'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(tenant_id=tenant_id)

        deleted_count, _ = queryset.delete()
        return Response({'deleted': deleted_count}, status=status.HTTP_200_OK)

    def perform_destroy(self, instance):
        snapshot = _serialize_route(instance)
        tenant = instance.tenant
        entity_id = instance.id
        response = super().perform_destroy(instance)
        _log_audit_event(
            tenant=tenant,
            actor=self.request.user,
            event_type='route.deleted',
            entity_type='route',
            entity_id=entity_id,
            changes=_format_deleted_changes(snapshot),
        )
        return response

    def _resolve_cycle_days(self, client) -> int | None:
        label_val = self._parse_cycle_from_label(getattr(client, 'contact_days_label', ''))
        if label_val:
            return label_val
        if client.contact_reminder_days and client.contact_reminder_days > 0:
            return int(client.contact_reminder_days)
        return None

    def _annotate_with_latest_contacts(self, queryset):
        latest_records = CallRecord.objects.filter(client=OuterRef('pk')).order_by('-contact_date', '-id')
        return queryset.annotate(
            last_contact_date=Subquery(latest_records.values('contact_date')[:1]),
            recorded_next_contact=Subquery(latest_records.values('next_contact_at')[:1]),
        )

    def _export_plan_excel(self, selected_date: date_cls, entries: list[dict]):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Plan"
        headers = [
            "Data", "Klient", "NIP", "Miasto", "Handlowiec", "Cykl (dni)",
            "Ostatni kontakt", "Planowany kontakt",
        ]
        sheet.append(headers)
        for entry in entries:
            sheet.append([
                entry['due_date'],
                entry['name'],
                entry['nip'],
                entry['city'],
                entry['salesman_name'] or "-",
                entry['cycle_days'],
                entry['last_contact_date'] or "",
                entry['due_date'],
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"plan_kontaktow_{selected_date.isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='lookup-by-nip')
    def lookup_by_nip(self, request):
        nip = request.query_params.get('nip', '').strip()
        if not nip:
            return Response({'detail': 'Podaj NIP.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            # Try scraper first
            data = lookup_client_by_nip_scraper(nip)
        except RegonAPIError as exc:
            # Fallback to API if scraper fails
            try:
                data = lookup_client_by_nip(nip)
            except RegonAPIError as api_exc:
                return Response({'detail': str(api_exc)}, status=status.HTTP_400_BAD_REQUEST)

        if not data:
            return Response({'detail': 'Nie znaleziono firmy w REGON.'}, status=status.HTTP_404_NOT_FOUND)

        return Response({
            'name': data.get('name') or '',
            'nip': data.get('nip') or '',
            'city': data.get('city') or '',
            'postal_code': data.get('postal_code') or '',
            'street': data.get('street') or '',
            'regon': data.get('regon') or '',
        })

    def _resolve_contact_cycle_tenant(self, request):
        tenant = getattr(request.user, 'tenant', None)
        tenant_id_param = request.query_params.get('tenant_id') or request.data.get('tenant_id')
        if tenant_id_param not in (None, '', 'null'):
            try:
                tenant_candidate = Tenant.objects.get(pk=int(tenant_id_param))
            except (Tenant.DoesNotExist, ValueError, TypeError):
                raise serializers.ValidationError({'tenant_id': 'Nieprawidłowy identyfikator tenanta.'})

            allowed = False
            if request.user.role == UserRole.ADMIN:
                allowed = True
            elif request.user.role == UserRole.MANAGER and tenant and tenant.pk == tenant_candidate.pk:
                allowed = True

            if not allowed:
                raise PermissionDenied('Brak uprawnień do modyfikacji wskazanego tenanta.')

            tenant = tenant_candidate

        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant_id': 'Brak zdefiniowanego tenanta.'})
        return tenant

    @action(detail=False, methods=['get', 'patch'], url_path='contact-cycle-start', permission_classes=[permissions.IsAuthenticated])
    def contact_cycle_start(self, request):
        tenant = self._resolve_contact_cycle_tenant(request)

        if request.method.lower() == 'get':
            return Response(
                {
                    'start_date': tenant.contact_cycle_start_date.isoformat() if tenant.contact_cycle_start_date else None,
                    'tenant_id': tenant.pk,
                }
            )

        if request.user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do zmiany dnia zerowego.')

        start_date_value = request.data.get('start_date')
        if start_date_value in (None, '', 'null'):
            tenant.contact_cycle_start_date = None
        else:
            parsed = parse_date(str(start_date_value))
            if not parsed:
                raise serializers.ValidationError({'start_date': 'Nieprawidłowy format daty. Użyj RRRR-MM-DD.'})
            tenant.contact_cycle_start_date = parsed

        tenant.save(update_fields=['contact_cycle_start_date'])
        return Response(
            {
                'start_date': tenant.contact_cycle_start_date.isoformat() if tenant.contact_cycle_start_date else None,
                'tenant_id': tenant.pk,
            }
        )

    @action(detail=True, methods=['get'], url_path='call-history')
    def call_history(self, request, pk=None):
        client = self.get_object()
        try:
            limit = max(1, min(50, int(request.query_params.get('limit', 5))))
        except (TypeError, ValueError):
            limit = 5
        try:
            offset = max(0, int(request.query_params.get('offset', 0)))
        except (TypeError, ValueError):
            offset = 0

        queryset = (
            CallRecord.objects.filter(client=client)
            .select_related('handler')
            .order_by('-contact_date', '-id')
        )
        total_count = queryset.count()
        records = queryset[offset : offset + limit]
        serializer = CallRecordSerializer(records, many=True)
        return Response(
            {
                'count': total_count,
                'limit': limit,
                'offset': offset,
                'results': serializer.data,
            }
        )

    @action(detail=True, methods=['post'], url_path='contact-plan/complete')
    def mark_contact_completed(self, request, pk=None):
        client = self.get_object()
        data = request.data or {}
        contact_date_value = data.get('contact_date')
        parsed_contact_date = parse_date(contact_date_value) if contact_date_value else None
        contact_date = parsed_contact_date or timezone.localdate()
        next_contact_value = data.get('next_contact_at')
        parsed_next_contact = parse_date(next_contact_value) if next_contact_value else None

        outcome = data.get('outcome', '')
        current_comment = data.get('current_comment', '')
        approval_reason = data.get('approval_reason', '')

        handler = request.user
        handler_id = data.get('handler')
        if handler_id and request.user.role in {UserRole.ADMIN, UserRole.MANAGER}:
            try:
                target = User.objects.filter(pk=int(handler_id)).first()
            except (TypeError, ValueError):
                target = None
            if target:
                handler = target

        cycle_days = self._resolve_cycle_days(client)
        auto_next_contact = parsed_next_contact
        if not auto_next_contact and cycle_days:
            candidate = contact_date + timedelta(days=cycle_days)
            auto_next_contact = shift_to_business_day(candidate)

        record = CallRecord.objects.create(
            tenant=client.tenant,
            client=client,
            handler=handler,
            contact_date=contact_date,
            contact_time=timezone.localtime().time(),
            next_contact_at=auto_next_contact,
            outcome=outcome,
            current_comment=current_comment,
        )
        approval_required = _create_contact_next_date_request_if_needed(record, cycle_days, reason=approval_reason)
        serializer = CallRecordSerializer(record)
        data = serializer.data
        data['approval_required'] = approval_required
        return Response(data, status=status.HTTP_201_CREATED)


class ClientDeletionRequestViewSet(TenantScopedViewSet):
    queryset = ClientDeletionRequest.objects.select_related('tenant', 'client', 'requested_by', 'reviewed_by')
    serializer_class = ClientDeletionRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role in {UserRole.ADMIN, UserRole.MANAGER}:
            return queryset
        return queryset.none()

    def get_queryset(self):
        queryset = super().get_queryset()
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        return queryset

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def approve(self, request, pk=None):
        deletion_request = self.get_object()
        if deletion_request.status != ClientDeletionRequest.Status.PENDING:
            return Response({'detail': 'Wniosek został już rozpatrzony.'}, status=status.HTTP_400_BAD_REQUEST)

        client = deletion_request.client
        deletion_request.status = ClientDeletionRequest.Status.APPROVED
        deletion_request.reviewed_by = request.user
        deletion_request.reviewed_at = timezone.now()
        deletion_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'updated_at'])

        client.status = Client.Status.DELETED
        client.deleted_at = timezone.now()
        client.save(update_fields=['status', 'deleted_at', 'updated_at'])

        _log_audit_event(
            tenant=client.tenant,
            actor=request.user,
            event_type='client.deletion_approved',
            entity_type='client',
            entity_id=client.id,
            changes={'status': {'from': Client.Status.ACTIVE, 'to': Client.Status.DELETED}},
        )

        serializer = self.get_serializer(deletion_request, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def reject(self, request, pk=None):
        deletion_request = self.get_object()
        if deletion_request.status != ClientDeletionRequest.Status.PENDING:
            return Response({'detail': 'Wniosek został już rozpatrzony.'}, status=status.HTTP_400_BAD_REQUEST)

        deletion_request.status = ClientDeletionRequest.Status.REJECTED
        deletion_request.reviewed_by = request.user
        deletion_request.reviewed_at = timezone.now()
        deletion_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'updated_at'])

        _log_audit_event(
            tenant=deletion_request.tenant,
            actor=request.user,
            event_type='client.deletion_rejected',
            entity_type='client',
            entity_id=deletion_request.client_id,
            changes={'deletion_request': {'from': ClientDeletionRequest.Status.PENDING, 'to': ClientDeletionRequest.Status.REJECTED}},
        )

        serializer = self.get_serializer(deletion_request, context={'request': request})
        return Response(serializer.data)


class TenantContactCycleCompatView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, tenant_id: int, *args, **kwargs):
        tenant = Tenant.objects.filter(pk=tenant_id).first()
        if not tenant:
            return Response({'detail': 'Tenant nie istnieje.'}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do zmiany dnia zerowego.')

        start_date_value = request.data.get('start_date')
        if start_date_value in (None, '', 'null'):
            tenant.contact_cycle_start_date = None
        else:
            parsed = parse_date(str(start_date_value))
            if not parsed:
                raise serializers.ValidationError({'start_date': 'Nieprawidłowy format daty. Użyj RRRR-MM-DD.'})
            tenant.contact_cycle_start_date = parsed

        tenant.save(update_fields=['contact_cycle_start_date'])
        return Response(
            {
                'start_date': tenant.contact_cycle_start_date.isoformat() if tenant.contact_cycle_start_date else None,
            }
        )

    @action(detail=False, methods=['get'], url_path='contact-stats')
    def contact_stats(self, request):
        date_param = request.query_params.get('date')
        target_date = parse_date(date_param) if date_param else timezone.localdate()
        user = request.user

        # Zaplanowane na dzień (z contact_plan) - z podziałem na handlowców
        queryset = self._annotate_with_latest_contacts(self.filter_queryset(self.get_queryset()))
        completion_subquery = CallRecord.objects.filter(client=OuterRef('pk'), contact_date=target_date)
        queryset = queryset.annotate(completed_on_selected=Exists(completion_subquery))
        
        scheduled_by_salesman = {}
        for client in queryset:
            cycle_days = self._resolve_cycle_days(client)
            if not cycle_days:
                continue
            created_date = (client.created_at.date() if client.created_at else target_date)
            client_tenant = getattr(client, 'tenant', None)
            salesman = getattr(client, 'salesman', None)
            salesman_cycle_start = getattr(salesman, 'contact_cycle_start_date', None) if salesman else None
            tenant_cycle_start = getattr(client_tenant, 'contact_cycle_start_date', None)
            client_cycle_start = salesman_cycle_start or tenant_cycle_start
            context = PlanComputationContext(
                created_date=created_date,
                cycle_days=cycle_days,
                last_contact_date=getattr(client, 'last_contact_date', None),
                recorded_next_date=getattr(client, 'recorded_next_contact', None),
                cycle_start_date=client_cycle_start,
            )
            result = compute_due_date(context, target_date)
            if result and result.due_date == target_date:
                salesman_name = self._format_salesman(salesman) or 'Brak handlowca'
                # Dla handlowca pokaż tylko jego dane
                if user.role == UserRole.REP and salesman_name != self._format_salesman(user):
                    continue
                scheduled_by_salesman[salesman_name] = scheduled_by_salesman.get(salesman_name, 0) + 1

        # Wykonane na dzień - z podziałem na handlowców
        completed_qs = self.filter_queryset(CallRecord.objects.filter(contact_date=target_date))
        completed_qs = completed_qs.select_related('handler')
        completed_by_salesman = {}
        for record in completed_qs:
            handler = record.handler
            salesman_name = f"{handler.first_name} {handler.last_name}".strip() if handler else ''
            if not salesman_name and handler:
                salesman_name = handler.username or 'Brak handlowca'
            # Dla handlowca pokaż tylko jego dane
            if user.role == UserRole.REP and handler and handler.id != user.id:
                continue
            completed_by_salesman[salesman_name] = completed_by_salesman.get(salesman_name, 0) + 1

        # Połącz wszystkie nazwy handlowców
        all_salesmen = set(scheduled_by_salesman.keys()) | set(completed_by_salesman.keys())
        
        stats_by_salesman = []
        for salesman_name in sorted(all_salesmen):
            stats_by_salesman.append({
                'salesman': salesman_name,
                'scheduled': scheduled_by_salesman.get(salesman_name, 0),
                'completed': completed_by_salesman.get(salesman_name, 0),
            })

        return Response({
            'by_salesman': stats_by_salesman,
            'total_scheduled': sum(scheduled_by_salesman.values()),
            'total_completed': sum(completed_by_salesman.values()),
        })

    @action(detail=True, methods=['get'], url_path='call-history')
    def call_history(self, request, pk=None):
        client = self.get_object()
        try:
            limit = max(1, min(50, int(request.query_params.get('limit', 5))))
        except (TypeError, ValueError):
            limit = 5
        try:
            offset = max(0, int(request.query_params.get('offset', 0)))
        except (TypeError, ValueError):
            offset = 0

        queryset = CallRecord.objects.filter(client=client).select_related('handler').order_by('-contact_date', '-id')
        total_count = queryset.count()
        records = queryset[offset: offset + limit]
        serializer = CallRecordSerializer(records, many=True)
        return Response({
            'count': total_count,
            'limit': limit,
            'offset': offset,
            'results': serializer.data,
        })


class VisitViewSet(TenantScopedViewSet):
    queryset = Visit.objects.select_related('tenant', 'client', 'salesman').all()
    serializer_class = VisitSerializer

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role == UserRole.REP:
            return queryset.filter(salesman=user)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        tenant = getattr(user, 'tenant', None)
        tenant_id = self.request.data.get('tenant') or self.request.data.get('tenant_id')
        if not tenant and tenant_id:
            try:
                tenant = Tenant.objects.filter(pk=int(tenant_id)).first()
            except (TypeError, ValueError):
                tenant = None
        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant': 'Brak przypisanego tenant.'})

        salesman = user
        salesman_id = self.request.data.get('salesman')
        if salesman_id and user.role in {UserRole.ADMIN, UserRole.MANAGER}:
            try:
                target = User.objects.filter(pk=int(salesman_id)).first()
            except (TypeError, ValueError):
                target = None
            if target:
                salesman = target
        visit = serializer.save(tenant=tenant, salesman=salesman)
        if (visit.status or '').lower() != 'confirmed':
            visit.status = 'confirmed'
            visit.save(update_fields=['status'])
        _log_audit_event(
            tenant=visit.tenant,
            actor=user,
            event_type='visit.created',
            entity_type='visit',
            entity_id=visit.id,
            changes=_format_created_changes(_serialize_visit(visit)),
        )

    @staticmethod
    def _strip_gps_note(text: str | None) -> str:
        if not text:
            return ''
        import re

        return re.sub(r"\s*\[GPS:[^\]]+\]\s*$", "", text, flags=re.IGNORECASE).strip()

    def update(self, request, *args, **kwargs):
        user = request.user
        if user.role == UserRole.REP:
            return Response({'detail': 'Brak uprawnień do edycji wizyt.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    @action(
        detail=False,
        methods=['delete'],
        url_path='purge',
        permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin],
    )
    def purge_visits(self, request):
        tenant_param = request.query_params.get('tenant')
        queryset = self.get_queryset()
        if tenant_param:
            try:
                tenant_id = int(tenant_param)
            except (TypeError, ValueError):
                return Response({'detail': 'Nieprawidłowy identyfikator tenanta.'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(tenant_id=tenant_id)

        deleted_count, _ = queryset.delete()
        return Response({'deleted': deleted_count}, status=status.HTTP_200_OK)

    def partial_update(self, request, *args, **kwargs):
        user = request.user
        if user.role == UserRole.REP:
            return Response({'detail': 'Brak uprawnień do edycji wizyt.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def perform_update(self, serializer):
        instance = serializer.instance
        user = self.request.user

        before = {
            'client': instance.client_id,
            'salesman': instance.salesman_id,
            'planned_at': instance.planned_at.isoformat() if instance.planned_at else None,
            'comment': self._strip_gps_note(instance.comment),
        }

        # Block any GPS/location updates from the client
        serializer.validated_data.pop('latitude', None)
        serializer.validated_data.pop('longitude', None)
        serializer.validated_data.pop('location_accuracy', None)
        serializer.validated_data.pop('location_name', None)

        # Only manager/admin may reassign salesman
        if user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            serializer.validated_data.pop('salesman', None)

        updated = serializer.save()

        after = {
            'client': updated.client_id,
            'salesman': updated.salesman_id,
            'planned_at': updated.planned_at.isoformat() if updated.planned_at else None,
            'comment': self._strip_gps_note(updated.comment),
        }

        changes = _diff_dict(before, after)

        if changes:
            _log_audit_event(
                tenant=updated.tenant,
                actor=user,
                event_type='visit.updated',
                entity_type='visit',
                entity_id=updated.id,
                changes=changes,
            )

    def perform_destroy(self, instance):
        snapshot = _serialize_visit(instance)
        tenant = instance.tenant
        entity_id = instance.id
        response = super().perform_destroy(instance)
        _log_audit_event(
            tenant=tenant,
            actor=self.request.user,
            event_type='visit.deleted',
            entity_type='visit',
            entity_id=entity_id,
            changes=_format_deleted_changes(snapshot),
        )
        return response


class RoutePlanViewSet(TenantScopedViewSet):
    queryset = RoutePlan.objects.select_related('tenant', 'owner', 'approved_by').prefetch_related('stops')
    serializer_class = RoutePlanSerializer

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role == UserRole.REP:
            return queryset.filter(owner=user)
        return queryset

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        owner = params.get('owner')
        exact_date = params.get('date')
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        approval_status = params.get('approval_status')

        if owner:
            try:
                queryset = queryset.filter(owner_id=int(owner))
            except (TypeError, ValueError):
                queryset = queryset.none()

        if exact_date:
            queryset = queryset.filter(date=exact_date)
        else:
            parsed_from = self._parse_date(date_from)
            parsed_to = self._parse_date(date_to)
            if parsed_from:
                queryset = queryset.filter(date__gte=parsed_from)
            if parsed_to:
                queryset = queryset.filter(date__lte=parsed_to)

        if approval_status:
            queryset = queryset.filter(approval_status=approval_status)

        # Always return newest routes first so dashboard queries with `limit` grab latest status
        return queryset.order_by('-date', '-id')

    @staticmethod
    def _parse_date(value: str | None):
        if not value:
            return None
        return parse_date(str(value))

    def perform_create(self, serializer):
        tenant = self._resolve_tenant()
        owner = self._resolve_owner()
        route = serializer.save(tenant=tenant, owner=owner)
        self._mark_pending(route)
        _log_audit_event(
            tenant=route.tenant,
            actor=self.request.user,
            event_type='route.created',
            entity_type='route',
            entity_id=route.id,
            changes=_format_created_changes(_serialize_route(route)),
        )

    def perform_update(self, serializer):
        owner = self._resolve_owner(default=serializer.instance.owner)
        instance = serializer.instance
        before = _serialize_route(instance)
        route = serializer.save(owner=owner)
        if self.request.user.role == UserRole.REP:
            self._mark_pending(route)
        after = _serialize_route(route)
        changes = _diff_dict(before, after)
        if changes:
            _log_audit_event(
                tenant=route.tenant,
                actor=self.request.user,
                event_type='route.updated',
                entity_type='route',
                entity_id=route.id,
                changes=changes,
            )

    def _resolve_tenant(self):
        tenant = getattr(self.request.user, 'tenant', None)
        tenant_id = self.request.data.get('tenant') or self.request.data.get('tenant_id')
        if not tenant and tenant_id:
            try:
                tenant = Tenant.objects.filter(pk=int(tenant_id)).first()
            except (TypeError, ValueError):
                tenant = None
        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant': 'Brak przypisanego tenant.'})
        return tenant

    def _resolve_owner(self, default=None):
        user = self.request.user
        owner = default or user
        owner_id = self.request.data.get('owner') or self.request.data.get('owner_id')
        if owner_id and user.role in {UserRole.ADMIN, UserRole.MANAGER}:
            try:
                target = User.objects.filter(pk=int(owner_id)).first()
            except (TypeError, ValueError):
                target = None
            if target:
                owner = target
        return owner

    def _mark_pending(self, route: RoutePlan):
        # Once a route has been approved, keep its status unless a manager explicitly changes it
        if route.approval_status == RoutePlan.ApprovalStatus.APPROVED:
            return
        if route.approval_status == RoutePlan.ApprovalStatus.PENDING and not route.approved_by:
            return
        route.approval_status = RoutePlan.ApprovalStatus.PENDING
        route.approved_at = None
        route.approved_by = None
        route.save(update_fields=['approval_status', 'approved_at', 'approved_by'])

    def _set_approval(self, route: RoutePlan, status_value: str):
        route.approval_status = status_value
        route.approved_at = timezone.now()
        route.approved_by = self.request.user
        route.save(update_fields=['approval_status', 'approved_at', 'approved_by'])
        _log_audit_event(
            tenant=route.tenant,
            actor=self.request.user,
            event_type=f'route.{status_value}',
            entity_type='route',
            entity_id=route.id,
            changes=_diff_dict({}, {'approval_status': status_value, 'approved_by': route.approved_by_id}),
        )
        serializer = self.get_serializer(route)
        return Response(serializer.data)

    @action(
        detail=True,
        methods=['post'],
        permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin],
    )
    def approve(self, request, *args, **kwargs):
        route = self.get_object()
        return self._set_approval(route, RoutePlan.ApprovalStatus.APPROVED)

    @action(
        detail=True,
        methods=['post'],
        permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin],
    )
    def reject(self, request, *args, **kwargs):
        route = self.get_object()
        return self._set_approval(route, RoutePlan.ApprovalStatus.REJECTED)

    @action(
        detail=False,
        methods=['delete'],
        url_path='purge',
        permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin],
    )
    def purge_routes(self, request):
        tenant_param = request.query_params.get('tenant')
        queryset = self.get_queryset()
        if tenant_param:
            try:
                tenant_id = int(tenant_param)
            except (TypeError, ValueError):
                return Response({'detail': 'Nieprawidłowy identyfikator tenanta.'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(tenant_id=tenant_id)

        deleted_count, _ = queryset.delete()
        return Response({'deleted': deleted_count}, status=status.HTTP_200_OK)

    @staticmethod
    def _parse_date(value):
        if not value:
            return None
        return parse_date(str(value))

 
class TaskViewSet(TenantScopedViewSet):
    queryset = (
        Task.objects.select_related('tenant', 'client', 'created_by', 'assigned_to', 'completed_by')
        .prefetch_related('messages__author')
        .all()
    )
    serializer_class = TaskSerializer

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role == UserRole.REP:
            return queryset.filter(assigned_to=user)
        return queryset

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        status_param = params.get('status')
        assigned = params.get('assigned_to')
        client = params.get('client')
        due_before = params.get('due_before')
        due_after = params.get('due_after')
        search = params.get('search')
        overdue = params.get('overdue')

        if status_param:
            queryset = queryset.filter(status=status_param)

        if assigned:
            try:
                queryset = queryset.filter(assigned_to_id=int(assigned))
            except (TypeError, ValueError):
                queryset = queryset.none()

        if client:
            try:
                queryset = queryset.filter(client_id=int(client))
            except (TypeError, ValueError):
                queryset = queryset.none()

        if due_before:
            queryset = queryset.filter(due_date__lte=due_before)
        if due_after:
            queryset = queryset.filter(due_date__gte=due_after)

        if overdue in {'1', 'true', 'True'}:
            queryset = queryset.filter(due_date__lt=timezone.localdate(), status__in=[
                Task.Status.PENDING,
                Task.Status.IN_PROGRESS,
                Task.Status.AWAITING_REVIEW,
            ])

        if search:
            queryset = queryset.filter(title__icontains=search)

        return queryset.order_by('due_date', 'id')

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do tworzenia zadań.')
        tenant = self._resolve_tenant()
        assigned_to = self._resolve_assignee()
        client = self._resolve_client()
        task = serializer.save(tenant=tenant, created_by=user, assigned_to=assigned_to, client=client)
        broadcast_task_update(task)
        _log_audit_event(
            tenant=task.tenant,
            actor=user,
            event_type='task.created',
            entity_type='task',
            entity_id=task.id,
            changes=_format_created_changes(_serialize_task(task)),
        )

    def update(self, request, *args, **kwargs):
        if request.user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do edycji zadań.')
        return super().update(request, *args, **kwargs)

    def perform_update(self, serializer):
        assigned_to = self._resolve_assignee(default=serializer.instance.assigned_to)
        client = self._resolve_client(default=serializer.instance.client)
        instance = serializer.instance
        before = _serialize_task(instance)
        task = serializer.save(assigned_to=assigned_to, client=client)
        broadcast_task_update(task)
        after = _serialize_task(task)
        changes = _diff_dict(before, after)
        if changes:
            _log_audit_event(
                tenant=task.tenant,
                actor=self.request.user,
                event_type='task.updated',
                entity_type='task',
                entity_id=task.id,
                changes=changes,
            )

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do usuwania zadań.')
        task = self.get_object()
        snapshot = _serialize_task(task)
        response = super().destroy(request, *args, **kwargs)
        broadcast_task_update(task)
        _log_audit_event(
            tenant=task.tenant,
            actor=request.user,
            event_type='task.deleted',
            entity_type='task',
            entity_id=task.id,
            changes=_format_deleted_changes(snapshot),
        )
        return response

    @action(detail=True, methods=['post'], url_path='messages')
    def add_message(self, request, pk=None):
        task = self.get_object()
        user = request.user
        body = (request.data.get('body') or '').strip()
        if not body:
            return Response({'detail': 'Treść wiadomości nie może być pusta.'}, status=status.HTTP_400_BAD_REQUEST)
        if user.role == UserRole.REP and task.assigned_to_id != user.id:
            raise PermissionDenied('Nie możesz edytować tego zadania.')

        is_manager_reply = user.role in {UserRole.ADMIN, UserRole.MANAGER}
        message = TaskMessage.objects.create(
            task=task,
            author=user,
            body=body,
            is_completion=False,
            is_manager_reply=is_manager_reply,
        )

        if user.role == UserRole.REP and task.status == Task.Status.PENDING:
            task.status = Task.Status.IN_PROGRESS
            task.save(update_fields=['status', 'updated_at'])
        elif is_manager_reply and task.status == Task.Status.AWAITING_REVIEW:
            task.status = Task.Status.IN_PROGRESS
            task.completed_at = None
            task.completed_by = None
            task.save(update_fields=['status', 'completed_at', 'completed_by', 'updated_at'])

        due_date_value = request.data.get('due_date')
        if is_manager_reply and due_date_value:
            parsed = parse_date(due_date_value)
            if parsed:
                task.due_date = parsed
                task.save(update_fields=['due_date', 'updated_at'])

        broadcast_task_message(message)
        broadcast_task_update(task)
        _log_audit_event(
            tenant=task.tenant,
            actor=user,
            event_type='task.message',
            entity_type='task',
            entity_id=task.id,
            changes=_format_created_changes({'message': message.body, 'author': message.author_id}),
        )
        serializer = TaskMessageSerializer(message)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='complete')
    def mark_complete(self, request, pk=None):
        task = self.get_object()
        user = request.user
        if user.role != UserRole.REP or task.assigned_to_id != user.id:
            raise PermissionDenied('Tylko przypisany handlowiec może oznaczyć zadanie.')
        body = (request.data.get('body') or '').strip()
        if not body:
            return Response({'detail': 'Dodaj krótką informację o wykonaniu.'}, status=status.HTTP_400_BAD_REQUEST)

        message = TaskMessage.objects.create(
            task=task,
            author=user,
            body=body,
            is_completion=True,
            is_manager_reply=False,
        )
        task.status = Task.Status.AWAITING_REVIEW
        task.save(update_fields=['status', 'updated_at'])
        broadcast_task_message(message)
        broadcast_task_update(task)
        _log_audit_event(
            tenant=task.tenant,
            actor=user,
            event_type='task.completed_request',
            entity_type='task',
            entity_id=task.id,
            changes=_format_created_changes({'message': message.body, 'status': task.status}),
        )
        return Response(TaskSerializer(task).data)

    @action(detail=True, methods=['post'], url_path='confirm')
    def confirm_completion(self, request, pk=None):
        task = self.get_object()
        user = request.user
        if user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do potwierdzenia zadania.')
        body = (request.data.get('body') or '').strip()
        if body:
            message = TaskMessage.objects.create(
                task=task,
                author=user,
                body=body,
                is_completion=True,
                is_manager_reply=True,
            )
            broadcast_task_message(message)
        task.status = Task.Status.COMPLETED
        task.completed_at = timezone.now()
        task.completed_by = user
        task.save(update_fields=['status', 'completed_at', 'completed_by', 'updated_at'])
        broadcast_task_update(task)
        _log_audit_event(
            tenant=task.tenant,
            actor=user,
            event_type='task.confirmed',
            entity_type='task',
            entity_id=task.id,
            changes=_format_created_changes({'status': task.status, 'completed_by': task.completed_by_id}),
        )
        return Response(TaskSerializer(task).data)

    @action(detail=True, methods=['post'], url_path='reopen')
    def reopen(self, request, pk=None):
        task = self.get_object()
        user = request.user
        if user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do zmiany zadania.')
        task.status = Task.Status.PENDING
        task.completed_at = None
        task.completed_by = None
        task.save(update_fields=['status', 'completed_at', 'completed_by', 'updated_at'])
        broadcast_task_update(task)
        _log_audit_event(
            tenant=task.tenant,
            actor=user,
            event_type='task.reopened',
            entity_type='task',
            entity_id=task.id,
            changes=_format_created_changes({'status': task.status}),
        )
        return Response(TaskSerializer(task).data)

    def _resolve_tenant(self):
        tenant = getattr(self.request.user, 'tenant', None)
        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant': 'Brak przypisanego tenant.'})
        return tenant

    def _resolve_assignee(self, default=None):
        assignee_id = self.request.data.get('assigned_to')
        if not assignee_id:
            return default
        try:
            user = User.objects.get(pk=int(assignee_id))
        except (User.DoesNotExist, TypeError, ValueError):
            raise serializers.ValidationError({'assigned_to': 'Nieprawidłowy użytkownik.'})
        if user.role != UserRole.REP:
            raise serializers.ValidationError({'assigned_to': 'Zadania można przypisać tylko do handlowca.'})
        return user

    def _resolve_client(self, default=None):
        client_id = self.request.data.get('client')
        if not client_id:
            return default
        try:
            client = Client.objects.get(pk=int(client_id))
        except (Client.DoesNotExist, TypeError, ValueError):
            raise serializers.ValidationError({'client': 'Nieprawidłowy klient.'})
        return client


class BackupJobViewSet(viewsets.ModelViewSet):
    queryset = BackupJob.objects.select_related('tenant', 'created_by').all()
    serializer_class = BackupJobSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.role == UserRole.ADMIN:
            return queryset
        tenant = getattr(user, 'tenant', None)
        if not tenant:
            return queryset.none()
        return queryset.filter(tenant=tenant)

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            raise PermissionDenied('Brak uprawnień do tworzenia kopii zapasowej.')
        tenant = getattr(user, 'tenant', None)
        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant': 'Brak przypisanego tenant.'})

        existing = BackupJob.objects.filter(tenant=tenant)
        for job in existing:
            job.delete()

        job = BackupJob.objects.create(tenant=tenant, created_by=user)
        run_backup_job.delay(job.id)
        serializer = self.get_serializer(job)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(
        detail=True,
        methods=['get'],
        url_path='download',
        permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin],
    )
    def download(self, request, pk=None):
        job = self.get_object()
        if job.status != BackupJob.Status.SUCCESS:
            return Response(
                {'detail': 'Backup nie jest gotowy do pobrania.'},
                status=status.HTTP_409_CONFLICT,
            )
        try:
            dump_bytes = generate_pg_dump_bytes()
        except RuntimeError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        filename = f"backup-{job.created_at.strftime('%Y%m%d-%H%M%S')}.dump"
        response = StreamingHttpResponse(
            streaming_content=iter([dump_bytes]),
            content_type='application/octet-stream',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = str(len(dump_bytes))
        return response


class CommentViewSet(TenantScopedViewSet):
    queryset = Comment.objects.select_related('tenant', 'client', 'author').all()
    serializer_class = CommentSerializer

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role == UserRole.REP:
            return queryset.filter(author=user)
        return queryset


class CallRecordViewSet(TenantScopedViewSet):
    queryset = CallRecord.objects.select_related('tenant', 'client', 'handler').all()
    serializer_class = CallRecordSerializer

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role == UserRole.REP:
            return queryset.filter(handler=user)
        return queryset

    @staticmethod
    def _parse_date_param(value):
        if not value:
            return None
        return parse_date(value)

    def perform_create(self, serializer):
        user = self.request.user
        tenant = getattr(user, 'tenant', None)
        tenant_id = self.request.data.get('tenant') or self.request.data.get('tenant_id')
        if not tenant and tenant_id:
            try:
                tenant = Tenant.objects.filter(pk=int(tenant_id)).first()
            except (TypeError, ValueError):
                tenant = None
        if not tenant:
            tenant = Tenant.objects.first()
        if not tenant:
            raise serializers.ValidationError({'tenant': 'Brak przypisanego tenant.'})

        handler = user
        handler_id = self.request.data.get('handler') or self.request.data.get('salesman')
        if handler_id and user.role in {UserRole.ADMIN, UserRole.MANAGER}:
            try:
                target = User.objects.filter(pk=int(handler_id)).first()
            except (TypeError, ValueError):
                target = None
            if target:
                handler = target
        record = serializer.save(tenant=tenant, handler=handler)
        _log_audit_event(
            tenant=record.tenant,
            actor=user,
            event_type='call.created',
            entity_type='call',
            entity_id=record.id,
            changes=_format_created_changes(_serialize_call_record(record)),
        )
        cycle_days = _resolve_client_cycle_days(record.client)
        _create_contact_next_date_request_if_needed(record, cycle_days)


    @action(
        detail=False,
        methods=['delete'],
        url_path='purge',
        permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin],
    )
    def purge_call_records(self, request):
        tenant_param = request.query_params.get('tenant')
        queryset = self.get_queryset()
        if tenant_param:
            try:
                tenant_id = int(tenant_param)
            except (TypeError, ValueError):
                return Response({'detail': 'Nieprawidłowy identyfikator tenanta.'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(tenant_id=tenant_id)

        deleted_count, _ = queryset.delete()
        return Response({'deleted': deleted_count}, status=status.HTTP_200_OK)

    def perform_update(self, serializer):
        instance = serializer.instance
        user = self.request.user

        before = {
            'client': instance.client_id,
            'handler': instance.handler_id,
            'contact_date': instance.contact_date.isoformat() if instance.contact_date else None,
            'next_contact_at': instance.next_contact_at.isoformat() if instance.next_contact_at else None,
            'outcome': instance.outcome,
            'current_comment': instance.current_comment,
        }

        if user.role not in {UserRole.ADMIN, UserRole.MANAGER}:
            serializer.validated_data.pop('handler', None)

        updated = serializer.save()

        after = {
            'client': updated.client_id,
            'handler': updated.handler_id,
            'contact_date': updated.contact_date.isoformat() if updated.contact_date else None,
            'next_contact_at': updated.next_contact_at.isoformat() if updated.next_contact_at else None,
            'outcome': updated.outcome,
            'current_comment': updated.current_comment,
        }

        changes = _diff_dict(before, after)

        if changes:
            _log_audit_event(
                tenant=updated.tenant,
                actor=user,
                event_type='call.updated',
                entity_type='call',
                entity_id=updated.id,
                changes=changes,
            )

    def perform_destroy(self, instance):
        snapshot = _serialize_call_record(instance)
        tenant = instance.tenant
        entity_id = instance.id
        response = super().perform_destroy(instance)
        _log_audit_event(
            tenant=tenant,
            actor=self.request.user,
            event_type='call.deleted',
            entity_type='call',
            entity_id=entity_id,
            changes=_format_deleted_changes(snapshot),
        )
        return response

    @action(detail=False, methods=['get'], url_path='export')
    def export_completed(self, request):
        query_date = request.query_params.get('date')
        target_date = self._parse_date_param(query_date) or timezone.localdate()

        queryset = self.filter_queryset(self.get_queryset()).filter(contact_date=target_date)
        queryset = queryset.select_related('client', 'handler')

        rows = []
        headers = [
            "Klient",
            "Handlowiec",
            "Data kontaktu",
            "Godzina",
            "Wynik",
            "Komentarz",
        ]
        for record in queryset:
            handler = record.handler
            handler_name = f"{handler.first_name} {handler.last_name}".strip() if handler else ''
            if not handler_name and handler:
                handler_name = handler.username or ''
            rows.append([
                getattr(record.client, 'name', '') or "",
                handler_name,
                record.contact_date.isoformat() if record.contact_date else "",
                record.contact_date.strftime("%H:%M") if record.contact_date else "",
                record.outcome or "",
                record.current_comment or "",
            ])

        export_format = request.query_params.get('format', 'xlsx').lower()
        filename_base = f"wykonane_kontakty_{target_date.isoformat()}"
        if export_format == 'pdf':
            return _build_pdf_response(
                f"{filename_base}.pdf",
                f"Wykonane kontakty – {target_date.isoformat()}",
                headers,
                rows,
                column_widths=[2.4, 1.6, 1.3, 0.9, 1.6, 2.4],
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Wykonane kontakty"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{filename_base}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related('tenant', 'actor').all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsManagerOrAdmin]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if user.role == UserRole.ADMIN:
            scoped = queryset
        elif user.role == UserRole.MANAGER and user.tenant_id:
            scoped = queryset.filter(tenant=user.tenant)
        else:
            scoped = queryset.none()

        params = self.request.query_params
        event_type = params.get('event_type')
        entity_type = params.get('entity_type')
        entity_id = params.get('entity_id')

        if event_type:
            scoped = scoped.filter(event_type=event_type)
        if entity_type:
            scoped = scoped.filter(entity_type=entity_type)
        if entity_id:
            try:
                scoped = scoped.filter(entity_id=int(entity_id))
            except (TypeError, ValueError):
                scoped = scoped.none()

        return scoped



class ImportJobViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet,
):
    queryset = ImportJob.objects.select_related('tenant', 'created_by').prefetch_related('records')
    serializer_class = ImportJobSerializer
    permission_classes = [IsManagerOrAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.role == UserRole.ADMIN:
            return queryset
        if user.role == UserRole.MANAGER and user.tenant_id:
            return queryset.filter(tenant=user.tenant)
        return queryset.none()

    def create(self, request, *args, **kwargs):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Brak pliku Excel.'}, status=status.HTTP_400_BAD_REQUEST)

        tenant = None
        tenant_id = request.data.get('tenant_id')
        if tenant_id:
            try:
                tenant = Tenant.objects.filter(id=int(tenant_id)).first()
            except (TypeError, ValueError):
                tenant = None
        if not tenant:
            tenant = getattr(request.user, 'tenant', None)

        if not tenant:
            return Response({'detail': 'Brak przypisanego tenant'}, status=status.HTTP_400_BAD_REQUEST)

        upload_content = upload.read()
        upload.seek(0)

        job = ImportJob.objects.create(
            tenant=tenant,
            created_by=request.user,
            upload=upload,
            status=ImportJob.Status.PENDING,
            upload_blob=upload_content,
        )
        process_import_job.delay(str(job.pk))
        serializer = self.get_serializer(job)
        headers = self.get_success_headers(serializer.data)
        _log_audit_event(
            tenant=job.tenant,
            actor=request.user,
            event_type='import.created',
            entity_type='import_job',
            entity_id=job.pk.int if hasattr(job.pk, 'int') else job.pk,
            changes=_format_created_changes({'filename': upload.name, 'status': job.status}),
        )
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        job = self.get_object()
        if job.status not in {ImportJob.Status.PENDING, ImportJob.Status.RUNNING}:
            return Response({'detail': 'Zadanie nie może zostać zatrzymane.'}, status=status.HTTP_400_BAD_REQUEST)
        if job.cancel_requested:
            return Response({'detail': 'Anulowanie już zgłoszone.'}, status=status.HTTP_202_ACCEPTED)
        job.cancel_requested = True
        job.save(update_fields=['cancel_requested'])
        serializer = self.get_serializer(job)
        _log_audit_event(
            tenant=job.tenant,
            actor=request.user,
            event_type='import.cancel_requested',
            entity_type='import_job',
            entity_id=job.pk.int if hasattr(job.pk, 'int') else job.pk,
            changes=_format_created_changes({'cancel_requested': True}),
        )
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)


class ContactNextDateRequestViewSet(TenantScopedViewSet):
    queryset = ContactNextDateRequest.objects.select_related(
        'tenant', 'client', 'call_record', 'requested_by', 'reviewed_by'
    ).all()
    serializer_class = ContactNextDateRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def filter_queryset_by_role(self, queryset):
        queryset = super().filter_queryset_by_role(queryset)
        user = self.request.user
        if user.role in {UserRole.ADMIN, UserRole.MANAGER}:
            return queryset
        return queryset.filter(requested_by=user)

    def get_queryset(self):
        queryset = super().get_queryset()
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        return queryset

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def approve(self, request, pk=None):
        req = self.get_object()
        if req.status != ContactNextDateRequest.Status.PENDING:
            return Response({'detail': 'Wniosek został już rozpatrzony.'}, status=status.HTTP_400_BAD_REQUEST)
        req.status = ContactNextDateRequest.Status.APPROVED
        req.reviewed_by = request.user
        req.reviewed_at = timezone.now()
        req.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])
        return Response(ContactNextDateRequestSerializer(req).data)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated, IsManagerOrAdmin])
    def reject(self, request, pk=None):
        req = self.get_object()
        if req.status != ContactNextDateRequest.Status.PENDING:
            return Response({'detail': 'Wniosek został już rozpatrzony.'}, status=status.HTTP_400_BAD_REQUEST)
        req.status = ContactNextDateRequest.Status.REJECTED
        req.reviewed_by = request.user
        req.reviewed_at = timezone.now()
        req.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])
        record = req.call_record
        if record.contact_date and req.cycle_days:
            standard_next = record.contact_date + timedelta(days=req.cycle_days)
            record.next_contact_at = standard_next
            record.save(update_fields=['next_contact_at'])
        return Response(ContactNextDateRequestSerializer(req).data)
